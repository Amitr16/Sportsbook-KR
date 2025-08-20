"""
Comprehensive tenant-filtered admin interface for sportsbook operators
"""

from flask import Blueprint, request, jsonify, session, render_template_string, redirect, send_file, make_response
import sqlite3
import json
from datetime import datetime, timedelta
from functools import wraps
import io
import csv
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

tenant_admin_bp = Blueprint('tenant_admin', __name__)

DATABASE_PATH = 'src/database/app.db'

def get_db_connection():
    """Get database connection"""
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def require_admin_auth(f):
    """Decorator to require admin authentication"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'operator_id' not in session:
            # If it's an API request, return JSON error
            if request.path.startswith('/api/') or request.headers.get('Content-Type') == 'application/json':
                return jsonify({
                    'success': False,
                    'error': 'Authentication required'
                }), 401
            # Otherwise redirect to login
            subdomain = kwargs.get('subdomain', '')
            return redirect(f'/admin/{subdomain}')
        return f(*args, **kwargs)
    return decorated_function

def get_current_operator():
    """Get current operator info from session"""
    if 'operator_id' not in session:
        return None
    
    conn = get_db_connection()
    operator = conn.execute("""
        SELECT id, login, sportsbook_name, subdomain, email, is_active, total_revenue, commission_rate
        FROM sportsbook_operators 
        WHERE id = ?
    """, (session['operator_id'],)).fetchone()
    conn.close()
    
    return dict(operator) if operator else None

@tenant_admin_bp.route('/api/admin/<subdomain>/events')
@require_admin_auth
def get_events(subdomain):
    """Get betting events (tenant-filtered)"""
    try:
        operator = get_current_operator()
        if not operator or operator['subdomain'] != subdomain:
            return jsonify({'success': False, 'error': 'Invalid operator'}), 403
        
        operator_id = operator['id']
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))
        sport = request.args.get('sport', '').strip()
        status = request.args.get('status', '').strip()
        
        conn = get_db_connection()
        
        # Build query with filters
        base_query = """
        FROM events e
        WHERE EXISTS (
            SELECT 1 FROM bets b 
            WHERE b.match_id = e.event_id 
            AND b.sportsbook_operator_id = ?
        )
        """
        params = [operator_id]
        
        if sport:
            base_query += " AND e.sport_name = ?"
            params.append(sport)
            
        if status:
            base_query += " AND e.status = ?"
            params.append(status)
        
        # Get total count
        total_count = conn.execute(f"SELECT COUNT(*) as count {base_query}", params).fetchone()['count']
        
        # Get paginated results
        offset = (page - 1) * per_page
        events = conn.execute(f"""
            SELECT 
                e.event_id, e.sport_name, e.home_team, e.away_team,
                e.event_date, e.status, e.score,
                COUNT(b.id) as bet_count,
                SUM(CASE WHEN b.status = 'pending' THEN b.stake ELSE 0 END) as total_stakes
            {base_query}
            GROUP BY e.event_id
            ORDER BY e.event_date DESC
            LIMIT ? OFFSET ?
        """, params + [per_page, offset]).fetchall()
        
        conn.close()
        
        return jsonify({
            'success': True,
            'events': [dict(event) for event in events],
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total_count,
                'pages': (total_count + per_page - 1) // per_page
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@tenant_admin_bp.route('/api/admin/<subdomain>/settle-bet', methods=['POST'])
@require_admin_auth
def settle_bet(subdomain):
    """Settle a bet manually (tenant-filtered)"""
    try:
        operator = get_current_operator()
        if not operator or operator['subdomain'] != subdomain:
            return jsonify({'success': False, 'error': 'Invalid operator'}), 403
        
        data = request.get_json()
        bet_id = data.get('bet_id')
        result = data.get('result')  # 'won' or 'lost'
        
        if not bet_id or result not in ['won', 'lost']:
            return jsonify({
                'success': False,
                'error': 'Invalid bet ID or result'
            }), 400
        
        operator_id = operator['id']
        conn = get_db_connection()
        
        # Verify bet belongs to this operator
        bet = conn.execute("""
            SELECT id, user_id, stake, potential_return, status
            FROM bets 
            WHERE id = ? AND sportsbook_operator_id = ? AND status = 'pending'
        """, (bet_id, operator_id)).fetchone()
        
        if not bet:
            conn.close()
            return jsonify({
                'success': False,
                'error': 'Bet not found or already settled'
            }), 404
        
        # Update bet status
        actual_return = bet['potential_return'] if result == 'won' else 0
        conn.execute("""
            UPDATE bets 
            SET status = ?, actual_return = ?, settled_at = ?
            WHERE id = ?
        """, (result, actual_return, datetime.utcnow(), bet_id))
        
        # Update user balance if won
        if result == 'won':
            conn.execute("""
                UPDATE users 
                SET balance = balance + ?
                WHERE id = ?
            """, (actual_return, bet['user_id']))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Bet settled as {result}'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@tenant_admin_bp.route('/api/admin/<subdomain>/user-details/<int:user_id>')
@require_admin_auth
def get_user_details(subdomain, user_id):
    """Get detailed user information (tenant-filtered)"""
    try:
        operator = get_current_operator()
        if not operator or operator['subdomain'] != subdomain:
            return jsonify({'success': False, 'error': 'Invalid operator'}), 403
        
        operator_id = operator['id']
        conn = get_db_connection()
        
        # Get user details
        user = conn.execute("""
            SELECT id, username, email, balance, created_at, last_login, is_active
            FROM users 
            WHERE id = ? AND sportsbook_operator_id = ?
        """, (user_id, operator_id)).fetchone()
        
        if not user:
            conn.close()
            return jsonify({
                'success': False,
                'error': 'User not found'
            }), 404
        
        # Get user's betting statistics
        stats = conn.execute("""
            SELECT 
                COUNT(*) as total_bets,
                COUNT(CASE WHEN status = 'won' THEN 1 END) as won_bets,
                COUNT(CASE WHEN status = 'lost' THEN 1 END) as lost_bets,
                COUNT(CASE WHEN status = 'pending' THEN 1 END) as pending_bets,
                SUM(stake) as total_staked,
                SUM(CASE WHEN status = 'won' THEN actual_return ELSE 0 END) as total_winnings,
                SUM(CASE WHEN status = 'lost' THEN stake ELSE 0 END) as total_losses
            FROM bets 
            WHERE user_id = ? AND sportsbook_operator_id = ?
        """, (user_id, operator_id)).fetchone()
        
        # Get recent bets
        recent_bets = conn.execute("""
            SELECT id, match_name, selection, stake, odds, potential_return, status, created_at
            FROM bets 
            WHERE user_id = ? AND sportsbook_operator_id = ?
            ORDER BY created_at DESC
            LIMIT 10
        """, (user_id, operator_id)).fetchall()
        
        conn.close()
        
        return jsonify({
            'success': True,
            'user': dict(user),
            'stats': dict(stats),
            'recent_bets': [dict(bet) for bet in recent_bets]
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@tenant_admin_bp.route('/api/admin/<subdomain>/update-user-balance', methods=['POST'])
@require_admin_auth
def update_user_balance(subdomain):
    """Update user balance (tenant-filtered)"""
    try:
        operator = get_current_operator()
        if not operator or operator['subdomain'] != subdomain:
            return jsonify({'success': False, 'error': 'Invalid operator'}), 403
        
        data = request.get_json()
        user_id = data.get('user_id')
        amount = data.get('amount')
        action = data.get('action')  # 'add' or 'subtract'
        
        if not user_id or not amount or action not in ['add', 'subtract']:
            return jsonify({
                'success': False,
                'error': 'Invalid parameters'
            }), 400
        
        operator_id = operator['id']
        conn = get_db_connection()
        
        # Verify user belongs to this operator
        user = conn.execute("""
            SELECT id, username, balance
            FROM users 
            WHERE id = ? AND sportsbook_operator_id = ?
        """, (user_id, operator_id)).fetchone()
        
        if not user:
            conn.close()
            return jsonify({
                'success': False,
                'error': 'User not found'
            }), 404
        
        # Calculate new balance
        current_balance = user['balance']
        if action == 'add':
            new_balance = current_balance + amount
        else:
            new_balance = max(0, current_balance - amount)  # Don't allow negative balance
        
        # Update balance
        conn.execute("""
            UPDATE users 
            SET balance = ?
            WHERE id = ?
        """, (new_balance, user_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Balance updated for {user["username"]}',
            'old_balance': current_balance,
            'new_balance': new_balance
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@tenant_admin_bp.route('/api/admin/<subdomain>/reports/generate', methods=['POST'])
@require_admin_auth
def generate_report(subdomain):
    """Generate custom reports (tenant-filtered)"""
    try:
        operator = get_current_operator()
        if not operator or operator['subdomain'] != subdomain:
            return jsonify({'success': False, 'error': 'Invalid operator'}), 403
        
        data = request.get_json()
        report_type = data.get('report_type')
        date_from = data.get('date_from')
        date_to = data.get('date_to')
        format_type = data.get('format', 'json')  # json, csv, pdf, excel
        
        if not report_type or not date_from or not date_to:
            return jsonify({
                'success': False,
                'error': 'Missing required parameters'
            }), 400
        
        operator_id = operator['id']
        conn = get_db_connection()
        
        # Generate report based on type
        if report_type == 'bets':
            query = """
            SELECT 
                b.id, b.match_name, b.selection, b.stake, b.odds, 
                b.potential_return, b.actual_return, b.status, b.created_at,
                u.username
            FROM bets b
            JOIN users u ON b.user_id = u.id
            WHERE b.sportsbook_operator_id = ?
            AND DATE(b.created_at) BETWEEN ? AND ?
            ORDER BY b.created_at DESC
            """
            data_rows = conn.execute(query, (operator_id, date_from, date_to)).fetchall()
            headers = ['ID', 'Match', 'Selection', 'Stake', 'Odds', 'Potential Return', 'Actual Return', 'Status', 'Date', 'User']
            
        elif report_type == 'users':
            query = """
            SELECT 
                u.id, u.username, u.email, u.balance, u.created_at, u.last_login,
                COUNT(b.id) as total_bets,
                SUM(CASE WHEN b.status = 'won' THEN b.actual_return ELSE 0 END) as total_winnings,
                SUM(CASE WHEN b.status = 'lost' THEN b.stake ELSE 0 END) as total_losses
            FROM users u
            LEFT JOIN bets b ON u.id = b.user_id AND b.sportsbook_operator_id = ?
            WHERE u.sportsbook_operator_id = ?
            AND DATE(u.created_at) BETWEEN ? AND ?
            GROUP BY u.id
            ORDER BY u.created_at DESC
            """
            data_rows = conn.execute(query, (operator_id, operator_id, date_from, date_to)).fetchall()
            headers = ['ID', 'Username', 'Email', 'Balance', 'Created', 'Last Login', 'Total Bets', 'Winnings', 'Losses']
            
        elif report_type == 'revenue':
            query = """
            SELECT 
                DATE(b.created_at) as date,
                COUNT(*) as total_bets,
                SUM(b.stake) as total_stakes,
                SUM(CASE WHEN b.status = 'won' THEN b.actual_return ELSE 0 END) as total_payouts,
                SUM(CASE WHEN b.status = 'lost' THEN b.stake ELSE 0 END) as revenue
            FROM bets b
            WHERE b.sportsbook_operator_id = ?
            AND DATE(b.created_at) BETWEEN ? AND ?
            GROUP BY DATE(b.created_at)
            ORDER BY date DESC
            """
            data_rows = conn.execute(query, (operator_id, date_from, date_to)).fetchall()
            headers = ['Date', 'Total Bets', 'Total Stakes', 'Total Payouts', 'Revenue']
            
        else:
            conn.close()
            return jsonify({
                'success': False,
                'error': 'Invalid report type'
            }), 400
        
        conn.close()
        
        # Convert to list of dictionaries
        report_data = []
        for row in data_rows:
            report_data.append(dict(row))
        
        # Return based on format
        if format_type == 'json':
            return jsonify({
                'success': True,
                'report_type': report_type,
                'date_range': f"{date_from} to {date_to}",
                'headers': headers,
                'data': report_data,
                'total_records': len(report_data)
            })
        
        elif format_type == 'csv':
            # Generate CSV
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            
            for row in data_rows:
                writer.writerow(list(row))
            
            # Create response
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'text/csv'
            response.headers['Content-Disposition'] = f'attachment; filename={report_type}_report_{date_from}_{date_to}.csv'
            return response
        
        elif format_type == 'excel':
            # Generate Excel
            output = io.BytesIO()
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = f"{report_type.title()} Report"
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Add data
            for row_idx, row in enumerate(data_rows, 2):
                for col_idx, value in enumerate(row, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(output)
            output.seek(0)
            
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = f'attachment; filename={report_type}_report_{date_from}_{date_to}.xlsx'
            return response
        
        else:
            return jsonify({
                'success': False,
                'error': 'Unsupported format'
            }), 400
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@tenant_admin_bp.route('/api/admin/<subdomain>/sports')
@require_admin_auth
def get_sports_list(subdomain):
    """Get list of sports with betting activity (tenant-filtered)"""
    try:
        operator = get_current_operator()
        if not operator or operator['subdomain'] != subdomain:
            return jsonify({'success': False, 'error': 'Invalid operator'}), 403
        
        operator_id = operator['id']
        conn = get_db_connection()
        
        sports = conn.execute("""
            SELECT 
                b.sport_name,
                COUNT(*) as bet_count,
                COUNT(DISTINCT b.match_id) as event_count,
                SUM(b.stake) as total_stakes,
                SUM(CASE WHEN b.status = 'won' THEN b.actual_return ELSE 0 END) as total_payouts
            FROM bets b
            WHERE b.sportsbook_operator_id = ?
            GROUP BY b.sport_name
            ORDER BY bet_count DESC
        """, (operator_id,)).fetchall()
        
        conn.close()
        
        return jsonify({
            'success': True,
            'sports': [dict(sport) for sport in sports]
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

